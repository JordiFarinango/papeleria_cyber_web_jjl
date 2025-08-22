from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from .models import Marca, Categoria, Proveedor, Producto, Cliente, Factura, DetalleFactura, IngresoInventario, Promocion
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils import timezone
from datetime import datetime
from django.db.models import F
from django.db import transaction
from django.http import HttpResponseNotAllowed

from django.utils.dateparse import parse_datetime

from decimal import Decimal

from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SEARCH_LIMIT = 8

#pagina antes del login
def inicio(request):
    return render(request, 'pagina_principal/inicio.html')

#vistas del administrador
def inicio_admin(request):
    return render(request, 'paginas_administrador/inicio_admin.html')

def nuevo_producto_admin(request):
    marcas = Marca.objects.all()
    categorias = Categoria.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        stock = request.POST.get('stock')
        precio = request.POST.get('precio')
        categoria_id = request.POST.get('categoria')
        marca_id = request.POST.get('marca')
        foto = request.FILES.get('foto')

        marca = get_object_or_404(Marca, id=marca_id)
        categoria = get_object_or_404(Categoria, id=categoria_id)

        if not nombre:
            messages.error(request,"Error al guardar el producto.")
            return redirect('nuevo_producto_admin')

        if nombre:
            guardar_producto = Producto(nombre=nombre, descripcion=descripcion, stock=stock, precio=precio, categoria=categoria, marca=marca, foto=foto)
            guardar_producto.save()
            messages.success(request, "Producto guardado correctamente.")
            return redirect('nuevo_producto_admin')
        
    return render(request, 'paginas_administrador/agregar/nuevo_producto.html', {
        'marcas_diccionario': marcas,
        'categorias_diccionario': categorias,
    })

def ver_productos_admin(request):
    productos = Producto.objects.all()
    marcas = Marca.objects.all()
    categorias = Categoria.objects.all()
    return render(request, 'paginas_administrador/ver/ver_productos.html',{
        'categorias': categorias,
        'marcas': marcas,
        'productos': productos
    })

def filtrar_productos_ver(request):
    cat_id = request.GET.get("cat")
    q = request.GET.get("q", "").strip()

    productos = Producto.objects.all()
    if cat_id:
        productos = productos.filter(categoria_id=cat_id)
    if q:
        productos = productos.filter(nombre__icontains=q)

    return render(request, "fragmentos/tabla_productos_ver.html", {"productos": productos})

def eliminar_producto(request):
    producto_id = request.POST.get("id_producto")
    producto_eli = get_object_or_404(Producto, id=producto_id)
    producto_eli.delete()
    return redirect('ver_productos_admin')

def editar_producto(request):
    id_producto = request.POST.get("id_producto_editar")
    nombre = request.POST.get("nombre")
    descripcion = request.POST.get("descripcion")
    stock = request.POST.get("stock")
    precio = request.POST.get("precio")
    marca_id = request.POST.get("marca_id")
    categoria_id = request.POST.get("categoria_id")
    foto = request.FILES.get("foto_url")

    if nombre:
        producto = get_object_or_404(Producto, id=id_producto)
        producto.nombre = nombre
        producto.descripcion = descripcion
        producto.stock = stock
        producto.precio = precio

        # 🔧 Aquí conviertes los IDs en instancias
        if marca_id:
            producto.marca = get_object_or_404(Marca, id=marca_id)
        if categoria_id:
            producto.categoria = get_object_or_404(Categoria, id=categoria_id)
        # Solo actualiza la foto si se cargó una nueva
        if foto:
            producto.foto = foto
        producto.save()
        messages.info(request, "Producto Actualizado Correctamente.")
        return redirect('ver_productos_admin')



def ingresar_inventario_admin(request):
    productos = Producto.objects.all()
    categorias = Categoria.objects.all()

    return render(request, 'paginas_administrador/agregar/ingresar_inventario.html',{
        'categorias': categorias,
        'productos': productos,
    })

def marcas_admin(request):
    marcas = Marca.objects.all()
    if request.method == 'POST': #si dio clic en guardar
        nombre = request.POST.get('nombre') #entonces guarda lo el contenido que haya escrito del cuadro del html con id nombre en esta varaible nombre

        if nombre: #verifica si la variable en verdad tiene algun valor
            nueva_marca = Marca(nombre=nombre) #quiero crear una nueva marca con este nombre (primer 'nombre' es del modelo Marca de models, y el segundo 'nombre' es de la variable creada arriba)
            nueva_marca.save() #se guarda la marca en la base de datos
            messages.success(request, "Marca guardado correctamente.")
            return redirect('marcas_admin') #cuando se guarde, redirigimos al usuarios a marcas.html
    return render(request, 'paginas_administrador/agregar/marcas.html',{
        'marcas': marcas
    })

def editar_marcas(request):
    marca_id = request.POST.get("marca_id")
    nombre = request.POST.get("nombre")

    if nombre:
        marca = get_object_or_404(Marca, id = marca_id)
        marca.nombre = nombre
        marca.save()
        messages.info(request, "Marca actualizada con éxito.")
        return redirect ("marcas_admin")

def categorias_admin(request):
    categorias = Categoria.objects.all()
    if request.method == 'POST':
        nombre = request.POST.get('nombre')

        if nombre:
            nueva_categoria = Categoria(nombre=nombre)
            nueva_categoria.save()
            messages.success(request, "Categoria guardado correctamente.")
            return redirect('categorias_admin')
    return render(request, 'paginas_administrador/agregar/categorias.html',{
        'categorias':categorias
    })

def editar_categoria(request):
    categoria_id = request.POST.get("id_categoria")
    nombre = (request.POST.get("nombre_cat") or "").strip()

    if not nombre:
        messages.error(request, "El nombre no puede estar vacío.")
        return redirect("categorias_admin")

    categoria = get_object_or_404(Categoria, id=categoria_id)
    categoria.nombre = nombre
    categoria.save()
    messages.info(request, "Categoría actualizada correctamente.")
    return redirect("categorias_admin")


def nuevo_proveedor_admin(request):
    proveedores = Proveedor.objects.all()
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        celular = request.POST.get('celular')
        cedula = request.POST.get('cedula')
        direccion = request.POST.get('direccion')

        if nombre and celular and cedula and direccion:
            nuevo_proveedor = Proveedor(nombre=nombre, celular=celular, ruc_cedula=cedula, direccion=direccion)
            nuevo_proveedor.save()
            messages.success(request, "Proveedor guardado correctamente")
            return redirect('nuevo_proveedor_admin')
    return render (request, 'paginas_administrador/agregar/nuevo_proveedor.html',{
        'proveedores' : proveedores
    })

def eliminar_proveedor(request):
    proveedor_id = request.POST.get("id_proveedor")
    proveedor = get_object_or_404(Proveedor, id=proveedor_id)
    proveedor.delete()
    messages.warning(request, "Proveedor eliminado correctamente.")
    return redirect('nuevo_proveedor_admin')

def editar_proveedor(request):
    id_proveedor = request.POST.get("id_proveedor_eli")
    nombre = request.POST.get("nombre")
    celular = request.POST.get("celular")
    ruc_cedula = request.POST.get("ruc_cedula")
    direccion = request.POST.get("direccion")

    if nombre:
        proveedor = get_object_or_404(Proveedor, id=id_proveedor)
        proveedor.nombre = nombre
        proveedor.celular = celular
        proveedor.ruc_cedula = ruc_cedula
        proveedor.direccion = direccion
        proveedor.save()
        messages.info(request,"Proveedor actualizado correctamente.")
        return redirect("nuevo_proveedor_admin")
   






def nuevo_cliente_admin(request):
    clientes = Cliente.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        cedula = request.POST.get('cedula')
        celular = request.POST.get('celular')
        direccion = request.POST.get('direccion')
        correo = request.POST.get('correo')
        provincia = request.POST.get('provincia')

        if Cliente.objects.filter(ruc_cedula=cedula).exists():
            messages.error(request, "Esta cédula ya existe.")
            return redirect('nuevo_cliente_admin')

        if nombre:
            nuevo_cliente = Cliente(nombre=nombre, ruc_cedula=cedula, direccion=direccion, correo=correo, celular=celular, provincia=provincia)
            nuevo_cliente.save()
            messages.success(request, "Cliente guardado correctamente.")
            return redirect('nuevo_cliente_admin')
        
    return render(request, 'paginas_administrador/agregar/nuevo_cliente.html',{
        'clientes_diccionario': clientes
    })
    

    
def facturar_admin(request):
    productos_iniciales = Producto.objects.all().order_by('nombre')[:SEARCH_LIMIT]
    return render(request, 'paginas_administrador/facturar.html', {
        'productos_iniciales': productos_iniciales
    })


def usuarios_admin(request):
    return render(request, 'paginas_administrador/agregar/usuarios.html')


def buscar_productos(request):
    q = (request.GET.get('q') or '').strip()
    if q:
        productos = Producto.objects.filter(nombre__icontains=q).order_by('nombre')[:8]
    else:
        # cuando no hay texto, mostramos los iniciales (coherente con la carga)
        productos = Producto.objects.all().order_by('nombre')[:10]

    html_filas = render_to_string('fragmentos/filas_productos.html', {'productos': productos})
    return JsonResponse({'html': html_filas})


def filtrar_productos_por_categoria(request, categoria_id):
    q = (request.GET.get('q') or '').strip()
    qs = Producto.objects.filter(categoria_id=categoria_id)
    if q:
        qs = qs.filter(nombre__icontains=q)
    productos = qs.order_by('nombre')[:50]  # limita para no traer 1000 filas
    return render(request, 'fragmentos/filtrar_productos_por_categoria.html', {
        'productos': productos
    })

@csrf_exempt
def actualizar_stock(request):
    if request.method == 'POST':
        try:
            datos = json.loads(request.body)
            producto_id = datos.get('producto_id')
            cantidad = int(datos.get('cantidad'))

            producto = Producto.objects.get(id=producto_id)
            producto.stock += cantidad
            producto.save()

            return JsonResponse({'exito': True, 'nuevo_stock': producto.stock})
        except Producto.DoesNotExist:
            return JsonResponse({'exito': False, 'error': 'Producto no encontrado'})
        except Exception as e:
            return JsonResponse({'exito': False, 'error': str(e)})

    return JsonResponse({'exito': False, 'error': 'Método no permitido'})


def eliminar_cliente(request):

    cliente_id = request.POST.get("cliente_id")
    cliente = get_object_or_404(Cliente, id=cliente_id)
    cliente.delete()
    messages.warning(request,"Cliente eliminado correctamente.")
    return redirect("nuevo_cliente_admin")

def editar_cliente(request):
    cliente_id = request.POST.get('modalClienteId')
    nombre = request.POST.get('nombre')
    cedula = request.POST.get('ruc_cedula')
    celular = request.POST.get('celular')
    direccion = request.POST.get('direccion')
    correo = request.POST.get('correo')
    provincia = request.POST.get('provincia')

    if nombre:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        cliente.nombre = nombre
        cliente.ruc_cedula = cedula
        cliente.celular = celular
        cliente.direccion = direccion
        cliente.correo = correo
        cliente.provincia = provincia
        cliente.save()
        messages.info(request, "Cliente actualizado correctamente.")

        return redirect('nuevo_cliente_admin')


def buscar_cliente_cedula(request):
    ced = request.GET.get('cedula', '').strip()
    if not ced:
        return JsonResponse({'found': False})
    try:
        cli = Cliente.objects.get(ruc_cedula=ced)
        return JsonResponse({
            'found': True,
            'cliente': {
                'id': cli.id,
                'nombre': cli.nombre,
                'celular': cli.celular,
                'direccion': cli.direccion,
                'correo': cli.correo,
                'provincia': cli.provincia,
            }
        })
    except Cliente.DoesNotExist:
        return JsonResponse({'found': False})
    



from django.http import FileResponse
from django.core.files.base import ContentFile

def factura_pdf(request, factura_id):
    factura = get_object_or_404(Factura, id=factura_id)

    # 1) Si ya existe el PDF sellado, lo servimos y listo.
    if factura.pdf_archivo:
        return FileResponse(factura.pdf_archivo.open('rb'), content_type='application/pdf')

    # 2) Generar por única vez con tu template actual.
    detalles = factura.detalles.select_related('producto').all()
    html = render_to_string('facturas/pdf.html', {
        'factura': factura,
        'detalles': detalles,
        'establecimiento': 'Papelería Mi Negocio',
    })

    result = BytesIO()
    pisa.CreatePDF(src=html, dest=result)
    pdf = result.getvalue()

    # 3) Guardar “sellado” en la factura y devolverlo.
    nombre_pdf = f'factura_{factura.id}.pdf'
    factura.pdf_archivo.save(nombre_pdf, ContentFile(pdf))
    factura.save()

    return FileResponse(BytesIO(pdf), content_type='application/pdf')

@csrf_exempt
def confirmar_venta(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'}, status=405)
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        cliente_payload = data.get('cliente')  # opcional: {'cedula','nombre','celular','direccion','correo','provincia'}

        if not items:
            return JsonResponse({'ok': False, 'error': 'Carrito vacío'})

        # 1) Resuelve cliente (opcional)
        cliente = None
        if cliente_payload:
            ced = (cliente_payload.get('cedula') or '').strip()
            nom = (cliente_payload.get('nombre') or '').strip()
            if ced:
                cliente = Cliente.objects.filter(ruc_cedula=ced).first()
                if not cliente and nom:
                    cliente = Cliente.objects.create(
                        nombre=nom,
                        ruc_cedula=ced,
                        celular=cliente_payload.get('celular') or '',
                        direccion=cliente_payload.get('direccion') or '',
                        correo=cliente_payload.get('correo') or '',
                        provincia=cliente_payload.get('provincia') or '',
                    )

        # 2) Crea la factura
        factura = Factura.objects.create(
            cliente=cliente,
            fecha=timezone.now(),
            total=Decimal('0.00'),
        )

        total = Decimal('0.00')

        # 3) Recorre items, descuenta stock y crea detalles
        for item in items:
            producto = Producto.objects.get(id=item['id'])
            cantidad = int(item['cantidad'])
            precio = Decimal(str(item['precio']))

            if producto.stock < cantidad:
                return JsonResponse({'ok': False, 'error': f'Sin stock para {producto.nombre}'})

            producto.stock -= cantidad
            producto.save()

            subtotal = precio * cantidad
            total += subtotal

            DetalleFactura.objects.create(
                factura=factura,
                producto=producto,
                cantidad=cantidad,
                precio=precio,
                subtotal=subtotal,
            )

        factura.total = total
        factura.save()
        messages.success(request, "Compra realizada con éxito.")


        return JsonResponse({'ok': True, 'factura_id': factura.id})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})
    

@csrf_exempt
def actualizar_stock_lote(request):
    if request.method != 'POST':
        return JsonResponse({'exito': False, 'error': 'Método no permitido'}, status=405)
    try:
        data = json.loads(request.body)
        items = data.get('items', [])  # [{producto_id, cantidad}, ...]

        if not items:
            return JsonResponse({'exito': False, 'error': 'Lote vacío'})

        actualizados = []

        for it in items:
            pid = it.get('producto_id')
            cant = int(it.get('cantidad') or 0)
            if not pid or cant <= 0:
                continue
            prod = Producto.objects.get(id=pid)
            prod.stock += cant
            prod.save()
            

            # registrar ingreso
            IngresoInventario.objects.create(
                producto=prod,
                cantidad_ingesada=cant,
                fecha_ingreso=timezone.now()
            )
            actualizados.append({'id': prod.id, 'nuevo_stock': prod.stock})

        if not actualizados:
            messages.error(request, "Nada valido para actualizar")
            return JsonResponse({'exito': False})
        
        messages.success(request, 'Stock actualizado correctamente.')
        return JsonResponse({'exito': True})

    except json.JSONDecodeError:
        messages.error(request, 'Formato de datos inválido.')
        return JsonResponse({'exito': False}, status=400)
    except Exception:
        messages.error(request, 'No se pudo actualizar el stock.')
        return JsonResponse({'exito': False}, status=500)


def historial_ventas(request):

    facturas = Factura.objects.none()  # vacío por defecto
    start = request.GET.get("start")
    end = request.GET.get("end")
    hoy_flag = request.GET.get("hoy")

    if hoy_flag == "1":
        hoy = timezone.localdate()  # respeta zona horaria
        facturas = Factura.objects.filter(fecha__date=hoy).order_by("-id")
    elif start or end:
        # Parseo seguro
        try:
            start_d = datetime.strptime(start, "%Y-%m-%d").date() if start else None
        except ValueError:
            start_d = None
        try:
            end_d = datetime.strptime(end, "%Y-%m-%d").date() if end else None
        except ValueError:
            end_d = None

        qs = Factura.objects.all()
        if start_d:
            qs = qs.filter(fecha__date__gte=start_d)
        if end_d:
            qs = qs.filter(fecha__date__lte=end_d)
        facturas = qs.order_by("-id")

    contexto = {
        "facturas": facturas,
        "start_value": start or "",
        "end_value": end or "",
        "aplico_filtro": (hoy_flag == "1") or bool(start or end),
    }
    return render(request, "paginas_administrador/ver/historial_ventas.html", contexto)

def factura_pdf_historial(request, pk):
    factura = get_object_or_404(
        Factura.objects.select_related("cliente").prefetch_related("detalles__producto"),
        pk=pk
    )
    contexto = {
        "factura": factura,
        "ahora": timezone.localtime(),  # por si quieres mostrar hora actual
    }
    return render(request, "facturas/pdf.html", contexto)

@transaction.atomic
def eliminar_venta(request, factura_id):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    factura = get_object_or_404(
        Factura.objects.prefetch_related("detalles__producto"),
        id=factura_id
    )

    # 1) Revertir stock de cada producto
    for det in factura.detalles.all():
        # Usamos F() para evitar condiciones de carrera
        Producto.objects.filter(id=det.producto_id).update(stock=F('stock') + det.cantidad)

    # 2) Borrar la factura (cascade borra detalles)
    factura.delete()
    messages.success(request, "Venta eliminada y stock revertido correctamente.")
    return redirect("historial_ventas")


def exportar_productos_excel(request):
    # Traer TODOS los productos (sin filtros/paginación)
    productos = (
        Producto.objects
        .select_related("marca", "categoria")
        .all()
        .order_by("nombre")
    )

    # Definir columnas EXACTAS (en orden)
    columnas = ["Nombre", "Marca", "Categoría", "Descripción", "Precio", "Stock"]

    # Crear workbook/hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados con formato
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for col_idx, titulo in enumerate(columnas, start=1):
        c = ws.cell(row=1, column=col_idx, value=titulo)
        c.font = header_font
        c.alignment = align_center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(titulo) + 2)

    # Escribir filas
    fila = 2
    for p in productos.iterator(chunk_size=1000):
        ws.cell(row=fila, column=1, value=p.nombre or "")
        ws.cell(row=fila, column=2, value=getattr(p.marca, "nombre", "") or "")
        ws.cell(row=fila, column=3, value=getattr(p.categoria, "nombre", "") or "")
        ws.cell(row=fila, column=4, value=p.descripcion or "")
        ws.cell(row=fila, column=5, value=float(p.precio) if p.precio is not None else None)
        ws.cell(row=fila, column=6, value=int(p.stock) if p.stock is not None else None)
        fila += 1

    # Formatos numéricos
    for r in range(2, fila):
        ws.cell(row=r, column=5).number_format = '#,##0.00'  # Precio
        ws.cell(row=r, column=6).number_format = '0'         # Stock

    # Preparar respuesta como archivo descargable
    now = timezone.localtime()
    filename = f"productos_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resp = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    return resp